import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .phishing_analyzer import PhishingAnalyzer
from .mitre_analyzer import MitreAttackAnalyzer

class SOCAnalyzer:
    def __init__(self):
        self.analyzer = PhishingAnalyzer()
        self.mitre_analyzer = MitreAttackAnalyzer()
        self.report_history = []
        
    def analyze_and_categorize(self, url):
        """
        Analyserer URL og kategoriserer risikonivå med tre nivåer
        """
        result = self.analyzer.check_url(url)
        
        # Legg til tidsstempel
        result['timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if result['status'] == 'completed':
            # Parse risk score
            risk_score = result.get('risk_score', 'N/A')
            if risk_score != 'N/A':
                positives, total = map(int, risk_score.split('/'))
                score_percent = (positives / total) * 100 if total > 0 else 0
                
                # Nye risikokategorier med oppdaterte grenser
                if score_percent < 3:  # 0-2 positive
                    result['risk_category'] = 'LAV'
                    result['action_required'] = 'Ingen umiddelbar handling nødvendig'
                elif score_percent < 10:  # 3-9 positive
                    result['risk_category'] = 'MEDIUM'
                    result['action_required'] = 'Undersøk nærmere - mulig falsk positiv'
                elif score_percent < 20:  # 10-19 positive
                    result['risk_category'] = 'HØY'
                    result['action_required'] = 'Umiddelbar handling påkrevd!'
                else:  # 20+ positive
                    result['risk_category'] = 'KRITISK'
                    result['action_required'] = 'KRITISK: Umiddelbar isolasjon og handling påkrevd!'
            else:
                result['risk_category'] = 'UKJENT'
                result['action_required'] = 'Kunne ikke bestemme risiko - manuell vurdering nødvendig'
        else:
            result['risk_category'] = 'FEIL'
            result['action_required'] = f'Analyse feilet - {result.get("error_message", "ukjent feil")}'
        
        # Debug utskrift før MITRE analyse
        print("\n=== DEBUG: MITRE Analysis Flow ===")
        print("1. Input URL:", url)
        
        # Legg til MITRE ATT&CK analyse
        analysis_input = {
            'url': url,
            'base_findings': result,
            'risk_category': result.get('risk_category', 'UKJENT')
        }
        print("2. Analysis Input:", json.dumps(analysis_input, indent=2))
        
        mitre_analysis = self.mitre_analyzer.analyze_threat(analysis_input)
        print("3. MITRE Analysis Result:", json.dumps(mitre_analysis, indent=2))
        
        # Kombiner resultatene
        result['mitre_analysis'] = {
            'techniques': mitre_analysis['identified_techniques'],
            'tactics': mitre_analysis['tactics'],
            'risk_score': mitre_analysis['risk_score']
        }
        print("4. Final Result Structure:", json.dumps({
            'mitre_analysis': result['mitre_analysis'],
            'mitre_details': result.get('mitre_details', {})
        }, indent=2))
        
        # Lagre resultatet i historikken
        self.report_history.append(result)
        return result
    
    def export_to_excel(self, filename="soc_reports.xlsx"):
        """
        Eksporterer analyserapporter til Excel med detaljert formatering
        """
        if not self.report_history:
            return "Ingen rapporter å eksportere"
            
        try:
            wb = Workbook()
            
            # Hovedark for alle analyser
            ws_main = wb.active
            ws_main.title = "URL Analysis"
            
            # Oppsummeringsark
            ws_summary = wb.create_sheet("Summary")
            ws_mitre = wb.create_sheet("MITRE Analysis")
            
            # ---- HOVEDARK FORMATERING ----
            headers = ['Timestamp', 'URL', 'Risk Category', 'Risk Score', 
                      'Action Required', 'VirusTotal Link', 'MITRE Score']
            
            # Formater headers
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Skriv og formater headers
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Farge-mapping for risikokategorier
            risk_colors = {
                'HØY': 'FF0000',      # Rød
                'MEDIUM': 'FFA500',    # Oransje
                'LAV': '90EE90',       # Lysegrønn
                'UKJENT': 'CCCCCC',    # Grå
                'FEIL': '000000'       # Sort
            }
            
            # Skriv data til hovedark
            for row, report in enumerate(self.report_history, 2):
                ws_main.cell(row=row, column=1, value=report.get('timestamp'))
                ws_main.cell(row=row, column=2, value=report.get('url'))
                
                risk_cell = ws_main.cell(row=row, column=3, value=report.get('risk_category'))
                risk_cell.fill = PatternFill(
                    start_color=risk_colors.get(report.get('risk_category', 'UKJENT')), 
                    end_color=risk_colors.get(report.get('risk_category', 'UKJENT')), 
                    fill_type="solid"
                )
                
                ws_main.cell(row=row, column=4, value=report.get('risk_score'))
                ws_main.cell(row=row, column=5, value=report.get('action_required'))
                ws_main.cell(row=row, column=6, value=report.get('permalink'))
                
                # Legg til MITRE score
                if 'mitre_analysis' in report:
                    ws_main.cell(row=row, column=7, value=report['mitre_analysis'].get('risk_score', 0))
            
            # ---- MITRE ANALYSIS ARK ----
            mitre_headers = ['URL', 'Techniques', 'Tactics', 'MITRE Risk Score']
            
            for col, header in enumerate(mitre_headers, 1):
                cell = ws_mitre.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for row, report in enumerate(self.report_history, 2):
                ws_mitre.cell(row=row, column=1, value=report.get('url'))
                
                if 'mitre_analysis' in report:
                    techniques = ', '.join(report['mitre_analysis'].get('techniques', []))
                    tactics = ', '.join(report['mitre_analysis'].get('tactics', []))
                    risk_score = report['mitre_analysis'].get('risk_score', 0)
                    
                    ws_mitre.cell(row=row, column=2, value=techniques)
                    ws_mitre.cell(row=row, column=3, value=tactics)
                    ws_mitre.cell(row=row, column=4, value=risk_score)
            
            # Juster kolonnebredder
            for ws in [ws_main, ws_summary, ws_mitre]:
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Lagre filen
            wb.save(filename)
            return f"Rapport eksportert til {filename}"
            
        except Exception as e:
            print(f"Feil ved eksport: {str(e)}")
            return f"Feil ved eksport: {str(e)}"
    
    def generate_summary(self):
        """
        Genererer en oppsummering med tre risikokategorier
        """
        if not self.report_history:
            return "Ingen analyser å oppsummere"
            
        summary = {
            'total_analyzed': len(self.report_history),
            'risk_distribution': {
                'HØY': {
                    'antall': 0,
                    'urls': []
                },
                'MEDIUM': {
                    'antall': 0,
                    'urls': []
                },
                'LAV': {
                    'antall': 0,
                    'urls': []
                },
                'UKJENT': {
                    'antall': 0,
                    'urls': []
                }
            },
            'latest_analysis': self.report_history[-1].get('timestamp', 'N/A')
        }
        
        # Samle URLer per kategori
        for report in self.report_history:
            category = report.get('risk_category', 'UKJENT')
            url = report.get('url', 'ukjent_url')
            summary['risk_distribution'][category]['antall'] += 1
            summary['risk_distribution'][category]['urls'].append({
                'url': url,
                'score': report.get('risk_score', 'N/A')
            })
            
        return summary
    
    def _extract_indicators(self, analysis_result):
        """Trekker ut relevante indikatorer for MITRE-analyse"""
        indicators = {
            'network_indicators': [],
            'host_indicators': [],
            'behavioral_patterns': []
        }
        
        try:
            # Nettverksindikatorer
            if 'url' in analysis_result:
                indicators['network_indicators'].append({
                    'type': 'url',
                    'value': analysis_result['url']
                })
                
            if 'ip_addresses' in analysis_result:
                indicators['network_indicators'].extend([
                    {'type': 'ip', 'value': ip} 
                    for ip in analysis_result['ip_addresses']
                ])
                
            # Host-indikatorer
            if 'downloaded_files' in analysis_result:
                indicators['host_indicators'].extend([
                    {'type': 'file', 'value': file_info}
                    for file_info in analysis_result['downloaded_files']
                ])
                
            # Atferdsmønstre
            if 'risk_category' in analysis_result:
                indicators['behavioral_patterns'].append({
                    'type': 'risk_level',
                    'value': analysis_result['risk_category']
                })
                
            if 'action_required' in analysis_result:
                indicators['behavioral_patterns'].append({
                    'type': 'recommended_action',
                    'value': analysis_result['action_required']
                })
                
        except Exception as e:
            print(f"Feil ved uttrekking av indikatorer: {str(e)}")
            
        return indicators

# Eksempel på bruk
if __name__ == "__main__":
    soc = SOCAnalyzer()
    
    # Test URLs
    test_urls = [
        "google.com",
        "eicar.org",
        "example.com"
    ]
    
    print("Starting SOC Analysis...")
    for url in test_urls:
        print(f"\nAnalyserer: {url}")
        result = soc.analyze_and_categorize(url)
        print(json.dumps(result, indent=2, ensure_ascii=False))
    
    print("\nGenererer oppsummering:")
    summary = soc.generate_summary()
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    
    print("\nEksporterer rapport...")
    print(soc.export_to_excel()) 